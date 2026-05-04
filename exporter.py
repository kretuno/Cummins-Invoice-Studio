from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from parser import ParseReport
from utils import format_countries, format_weights


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TOTAL_FILL = PatternFill("solid", fgColor="EAF4EA")
HEADER_FONT = Font(bold=True)


def export_report_to_excel(report: ParseReport, destination: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    notes = workbook.create_sheet("Notes")

    _build_summary_sheet(summary, report)
    _build_notes_sheet(notes, report)

    workbook.save(destination)


def _build_summary_sheet(sheet, report: ParseReport) -> None:
    headers = [
        "Invoice No",
        "Source File",
        "Positions",
        "Origin Countries",
        "Places",
        "Line Weights (kg)",
        "Net Weight (kg)",
        "Gross Weight (kg)",
        "Total USD",
    ]
    sheet.append(headers)

    for row in sheet[1]:
        row.font = HEADER_FONT
        row.fill = HEADER_FILL
        row.alignment = Alignment(horizontal="center", vertical="center")

    for record in report.invoices:
        sheet.append(
            [
                record.invoice_no,
                record.source_file,
                record.positions,
                format_countries(record.origin_countries),
                record.package_count if record.package_count is not None else "",
                format_weights(record.line_weights),
                record.net_weight,
                record.gross_weight if record.gross_weight is not None else "",
                record.total_usd,
            ]
        )

    total_row = sheet.max_row + 2
    sheet.cell(total_row, 4, "ALL Origin Countries")
    sheet.cell(total_row, 5, _format_all_origin_countries(report))
    sheet.cell(total_row + 1, 5, "TOTAL Places")
    sheet.cell(total_row + 1, 6, report.total_packages or "")
    sheet.cell(total_row + 2, 5, "TOTAL Net Weight")
    sheet.cell(total_row + 2, 7, report.total_weight)
    sheet.cell(total_row + 3, 5, "TOTAL Gross Weight")
    sheet.cell(total_row + 3, 8, report.total_gross_weight or "")
    sheet.cell(total_row + 4, 5, "TOTAL USD")
    sheet.cell(total_row + 4, 9, report.total_usd)

    for row_idx in (total_row, total_row + 1, total_row + 2, total_row + 3, total_row + 4):
        for col_idx in range(4, 10):
            cell = sheet.cell(row_idx, col_idx)
            cell.fill = TOTAL_FILL
            cell.font = HEADER_FONT

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=7, max_col=9):
        row[0].number_format = "0.000"
        row[1].number_format = "0.000"
        row[2].number_format = '#,##0.00'

    sheet.freeze_panes = "A2"
    _set_column_widths(
        sheet,
        {
            "A": 14,
            "B": 34,
            "C": 10,
            "D": 18,
            "E": 10,
            "F": 42,
            "G": 16,
            "H": 17,
            "I": 14,
        },
    )


def _build_notes_sheet(sheet, report: ParseReport) -> None:
    headers = ["Type", "Source / Invoice", "Details"]
    sheet.append(headers)
    for row in sheet[1]:
        row.font = HEADER_FONT
        row.fill = HEADER_FILL

    if not report.duplicates and not report.issues:
        sheet.append(["INFO", "-", "No duplicates or parsing notes."])

    for duplicate in report.duplicates:
        sheet.append(
            [
                "Duplicate",
                duplicate.invoice_no,
                f"Kept: {duplicate.kept_file}; Excluded: {duplicate.excluded_file}; {duplicate.reason}",
            ]
        )

    for issue in report.issues:
        sheet.append([issue.level, issue.source_file, issue.message])

    sheet.freeze_panes = "A2"
    _set_column_widths(sheet, {"A": 14, "B": 28, "C": 110})


def _set_column_widths(sheet, mapping: dict[str, int]) -> None:
    for column, width in mapping.items():
        sheet.column_dimensions[column].width = width


def _format_all_origin_countries(report: ParseReport) -> str:
    countries: list[str] = []
    for record in report.invoices:
        countries.extend(record.origin_countries)
    return format_countries(countries)
